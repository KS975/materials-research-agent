from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import datetime, timezone
import hashlib
import json
import math
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any
import zipfile

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency guard
    raise RuntimeError(
        "单位真实数据导入需要 openpyxl。请先执行 pip install -r requirements.txt"
    ) from exc


STAGE = "V0.2-company-real-data-import"
SCHEMA_VERSION = 1
LOCAL_PROJECT_ID_BASE = 930000

CANONICAL_FILES = {
    "samples": "样品.xlsx",
    "formula": "配方.xlsx",
    "performance": "测试性能.xlsx",
    "process": "工艺.xlsx",
    "conditions": "测试条件.xlsx",
    "service_performance": "服役性能.xlsx",
    "service_conditions": "服役条件.xlsx",
    "materials": "原料模板.xlsx",
}


class CompanyDataError(RuntimeError):
    pass


class CompanyDataValidationError(CompanyDataError):
    pass


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _atomic_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(tmp, path)


def _safe_extract(zip_path: Path, dest: Path) -> None:
    dest = dest.resolve()
    with zipfile.ZipFile(zip_path) as z:
        for info in z.infolist():
            name = info.filename.replace("\\", "/")
            if not name or name.endswith("/"):
                continue
            target = (dest / name).resolve()
            try:
                target.relative_to(dest)
            except ValueError as exc:
                raise CompanyDataValidationError(
                    f"ZIP 包含非法路径: {info.filename}"
                ) from exc
            target.parent.mkdir(parents=True, exist_ok=True)
            with z.open(info) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)


def _find_total_dir(extracted: Path) -> Path:
    candidates = [
        p for p in extracted.rglob("总库")
        if p.is_dir() and p.parent.name == "海科数据整理"
    ]
    if not candidates:
        raise CompanyDataValidationError(
            "未找到 海科数据整理/总库，无法确定 canonical 数据源"
        )
    if len(candidates) > 1:
        raise CompanyDataValidationError(
            "检测到多个 海科数据整理/总库，无法安全自动选择"
        )
    return candidates[0]


def _header(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _sample_name(value: Any) -> str:
    return _header(value)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_number(value: Any) -> float | None:
    if _is_blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        out = float(value)
        return out if math.isfinite(out) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        out = float(text)
        return out if math.isfinite(out) else None
    except ValueError:
        return None


def _read_first_sheet(path: Path) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        try:
            first = next(iterator)
        except StopIteration:
            return [], []
        headers = [_header(x) for x in first]
        rows: list[list[Any]] = []
        for row in iterator:
            values = list(row[: len(headers)])
            if any(not _is_blank(v) for v in values):
                rows.append(values)
        return headers, rows
    finally:
        wb.close()


def _row_mapping(
    headers: list[str],
    rows: list[list[Any]],
    *,
    key_name: str = "样品名称",
) -> dict[str, dict[str, Any]]:
    if not headers or headers[0] != key_name:
        raise CompanyDataValidationError(
            f"首列应为 {key_name!r}，实际为 {headers[0] if headers else '<empty>'!r}"
        )
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = _sample_name(row[0] if row else None)
        if not key:
            continue
        if key in out:
            raise CompanyDataValidationError(
                f"样品键重复: {key}"
            )
        record = {}
        for i, name in enumerate(headers):
            if not name:
                continue
            record[name] = row[i] if i < len(row) else None
        out[key] = record
    return out


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _fingerprint_files(total_dir: Path) -> dict[str, dict[str, Any]]:
    result = {}
    for key, filename in CANONICAL_FILES.items():
        path = total_dir / filename
        if not path.exists():
            raise CompanyDataValidationError(
                f"canonical 文件缺失: {path}"
            )
        result[key] = {
            "filename": filename,
            "sha256": sha256_file(path),
            "size_bytes": path.stat().st_size,
        }
    return result


def _subset_inventory(extracted: Path) -> list[dict[str, Any]]:
    root = next(
        (p for p in extracted.rglob("海科数据整理") if p.is_dir()),
        None,
    )
    if root is None:
        return []
    result = []
    for folder in sorted(
        [p for p in root.iterdir() if p.is_dir() and p.name != "总库"],
        key=lambda p: p.name.casefold(),
    ):
        sample_file = folder / "样品.xlsx"
        if not sample_file.exists():
            continue
        try:
            headers, rows = _read_first_sheet(sample_file)
            count = len(rows) if headers else 0
        except Exception:
            count = None
        result.append(
            {
                "name": folder.name,
                "sample_rows": count,
                "merged_into_canonical": False,
            }
        )
    return result


def import_company_archive(
    *,
    source_zip: str | Path,
    runtime_root: str | Path = ".runtime",
) -> dict[str, Any]:
    source_zip = Path(source_zip).expanduser().resolve()
    if not source_zip.exists():
        raise CompanyDataValidationError(
            f"source zip 不存在: {source_zip}"
        )
    if source_zip.suffix.lower() != ".zip":
        raise CompanyDataValidationError(
            "source 文件必须是 .zip"
        )

    runtime_root = Path(runtime_root)
    store_root = runtime_root / "company_data"
    archive_sha = sha256_file(source_zip)
    dataset_id = f"haike_{archive_sha[:12]}"
    import_dir = store_root / "imports" / dataset_id
    manifest_path = import_dir / "manifest.json"

    # Deterministic idempotency: same exact archive -> same import.
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        if manifest.get("source", {}).get("sha256") != archive_sha:
            raise CompanyDataValidationError(
                "dataset_id 冲突：manifest source hash 不一致"
            )
        _atomic_json(
            store_root / "current.json",
            {
                "dataset_id": dataset_id,
                "activated_at": utc_now_iso(),
                "source_sha256": archive_sha,
            },
        )
        return {
            "idempotent_replay": True,
            "manifest": manifest,
            "manifest_path": str(manifest_path),
        }

    with tempfile.TemporaryDirectory(prefix="company-data-") as tmp:
        extracted = Path(tmp) / "extracted"
        extracted.mkdir()
        _safe_extract(source_zip, extracted)

        total_dir = _find_total_dir(extracted)
        fingerprints = _fingerprint_files(total_dir)

        sample_h, sample_rows = _read_first_sheet(total_dir / CANONICAL_FILES["samples"])
        formula_h, formula_rows = _read_first_sheet(total_dir / CANONICAL_FILES["formula"])
        perf_h, perf_rows = _read_first_sheet(total_dir / CANONICAL_FILES["performance"])
        process_h, process_rows = _read_first_sheet(total_dir / CANONICAL_FILES["process"])
        cond_h, cond_rows = _read_first_sheet(total_dir / CANONICAL_FILES["conditions"])
        sp_h, sp_rows = _read_first_sheet(total_dir / CANONICAL_FILES["service_performance"])
        sc_h, sc_rows = _read_first_sheet(total_dir / CANONICAL_FILES["service_conditions"])
        mat_h, mat_rows = _read_first_sheet(total_dir / CANONICAL_FILES["materials"])

        sample_map = _row_mapping(sample_h, sample_rows)
        formula_map = _row_mapping(formula_h, formula_rows)
        perf_map = _row_mapping(perf_h, perf_rows)
        process_map = _row_mapping(process_h, process_rows)

        sample_keys = set(sample_map)
        warnings: list[str] = []
        for label, mapping in (
            ("配方", formula_map),
            ("测试性能", perf_map),
            ("工艺", process_map),
        ):
            if set(mapping) != sample_keys:
                missing = sorted(sample_keys - set(mapping))
                extra = sorted(set(mapping) - sample_keys)
                warnings.append(
                    f"{label} 与样品表键集合不完全一致: "
                    f"missing={len(missing)}, extra={len(extra)}"
                )

        material_ids = [_header(x) for x in formula_h[1:] if _header(x)]
        if len(set(material_ids)) != len(material_ids):
            raise CompanyDataValidationError(
                "配方表存在重复原料列"
            )

        performance_metrics = [
            _header(x) for x in perf_h[1:] if _header(x)
        ]
        if len(set(performance_metrics)) != len(performance_metrics):
            raise CompanyDataValidationError(
                "测试性能表存在重复性能列"
            )

        # Material template is an independent cross-check.
        material_template_names = {
            _header(row[0])
            for row in mat_rows
            if row and _header(row[0])
        }
        header_material_names = set(material_ids)
        if material_template_names != header_material_names:
            warnings.append(
                "原料模板与配方列集合不完全一致: "
                f"template_only={len(material_template_names-header_material_names)}, "
                f"formula_only={len(header_material_names-material_template_names)}"
            )

        product_counts: Counter[str] = Counter()
        product_formula_samples: Counter[str] = Counter()
        product_performance_samples: Counter[str] = Counter()
        product_metric_counts: dict[str, Counter[str]] = defaultdict(Counter)
        metric_counts: Counter[str] = Counter()
        metric_numeric_counts: Counter[str] = Counter()

        samples_out: list[dict[str, Any]] = []
        formulas_long: list[dict[str, Any]] = []
        performances_long: list[dict[str, Any]] = []
        process_out: list[dict[str, Any]] = []
        wide_rows: list[dict[str, Any]] = []

        ordered_sample_names = list(sample_map.keys())

        for idx, name in enumerate(ordered_sample_names, start=1):
            sample = sample_map[name]
            product_type = _header(sample.get("产品类型"))
            craft = _header(sample.get("工艺"))
            sample_key = f"HK{idx:06d}"

            product_counts[product_type] += 1

            formula = formula_map.get(name, {})
            perf = perf_map.get(name, {})
            proc = process_map.get(name, {})

            formula_values: dict[str, float] = {}
            for material_id in material_ids:
                value = _as_number(formula.get(material_id))
                if value is None:
                    continue
                formula_values[material_id] = value
                formulas_long.append(
                    {
                        "dataset_id": dataset_id,
                        "sample_key": sample_key,
                        "sample_name": name,
                        "product_type": product_type,
                        "material_id": material_id,
                        "amount": value,
                    }
                )
            if formula_values:
                product_formula_samples[product_type] += 1

            perf_values: dict[str, Any] = {}
            nonempty_perf = False
            for metric in performance_metrics:
                raw = perf.get(metric)
                if _is_blank(raw):
                    continue
                nonempty_perf = True
                number = _as_number(raw)
                perf_values[metric] = raw
                metric_counts[metric] += 1
                product_metric_counts[product_type][metric] += 1
                if number is not None:
                    metric_numeric_counts[metric] += 1
                performances_long.append(
                    {
                        "dataset_id": dataset_id,
                        "sample_key": sample_key,
                        "sample_name": name,
                        "product_type": product_type,
                        "metric": metric,
                        "value_raw": raw,
                        "value_numeric": number if number is not None else "",
                        "is_numeric": number is not None,
                    }
                )
            if nonempty_perf:
                product_performance_samples[product_type] += 1

            login_category = _header(proc.get("LOGINCATEGORY"))
            task_category = _header(proc.get("TASKCATEGORY"))

            samples_out.append(
                {
                    "dataset_id": dataset_id,
                    "sample_key": sample_key,
                    "sample_name": name,
                    "product_type": product_type,
                    "craft": craft,
                }
            )
            process_out.append(
                {
                    "dataset_id": dataset_id,
                    "sample_key": sample_key,
                    "sample_name": name,
                    "product_type": product_type,
                    "LOGINCATEGORY": login_category,
                    "TASKCATEGORY": task_category,
                    "classification": "workflow_metadata_not_material_process_parameter",
                }
            )

            wide = {
                "dataset_id": dataset_id,
                "sample_key": sample_key,
                "sample_name": name,
                "product_type": product_type,
                "craft": craft,
                "meta::LOGINCATEGORY": login_category,
                "meta::TASKCATEGORY": task_category,
            }
            for material_id in material_ids:
                wide[f"formula::{material_id}"] = (
                    formula_values.get(material_id, "")
                )
            for metric in performance_metrics:
                wide[f"performance::{metric}"] = (
                    perf_values.get(metric, "")
                )
            wide_rows.append(wide)

        products_sorted = sorted(
            product_counts,
            key=lambda x: (x.casefold(), x),
        )
        product_summaries = []
        for offset, product in enumerate(products_sorted, start=1):
            counts = product_metric_counts[product]
            top_metrics = [
                {"metric": metric, "nonempty_count": int(count)}
                for metric, count in counts.most_common()
            ]
            product_summaries.append(
                {
                    "local_project_id": LOCAL_PROJECT_ID_BASE + offset,
                    "product_type": product,
                    "sample_count": int(product_counts[product]),
                    "formula_present_samples": int(
                        product_formula_samples[product]
                    ),
                    "performance_present_samples": int(
                        product_performance_samples[product]
                    ),
                    "performance_coverage": top_metrics,
                }
            )

        top_products = [
            {
                "product_type": product,
                "sample_count": int(count),
                "local_project_id": next(
                    x["local_project_id"]
                    for x in product_summaries
                    if x["product_type"] == product
                ),
            }
            for product, count in product_counts.most_common(20)
        ]
        top_metrics = [
            {
                "metric": metric,
                "nonempty_count": int(count),
                "numeric_count": int(metric_numeric_counts[metric]),
            }
            for metric, count in metric_counts.most_common()
        ]

        canonical_rel = "海科数据整理/总库"
        subset_inventory = _subset_inventory(extracted)

        summary = {
            "samples": len(samples_out),
            "products": len(product_counts),
            "materials": len(material_ids),
            "performance_metrics": len(performance_metrics),
            "formula_present_samples": sum(
                product_formula_samples.values()
            ),
            "performance_present_samples": sum(
                product_performance_samples.values()
            ),
            "workflow_metadata_rows": len(process_out),
            "material_process_parameter_rows": 0,
            "explicit_test_condition_rows": len(cond_rows),
            "service_performance_rows": len(sp_rows),
            "service_condition_rows": len(sc_rows),
            "named_subset_directories": len(subset_inventory),
        }

        # These are intentionally blocking facts, not warnings to be silently ignored.
        safety = {
            "canonical_source": canonical_rel,
            "raw_archive_copied_into_runtime": False,
            "workflow_fields_are_process_features": False,
            "test_conditions_available": bool(cond_rows),
            "material_process_parameters_available": False,
            "official_model_allowed_from_import_alone": False,
            "reason": (
                "测试条件表未提供显式记录，且工艺表只有工作流分类字段；"
                "导入后可查询/统计/Reality Check，但不得绕过 Modeling Gate "
                "直接用于正式逆向设计或闭环 BO。"
            ),
        }

        manifest = {
            "stage": STAGE,
            "schema_version": SCHEMA_VERSION,
            "dataset_id": dataset_id,
            "imported_at": utc_now_iso(),
            "source": {
                "archive_name": source_zip.name,
                "sha256": archive_sha,
                "size_bytes": source_zip.stat().st_size,
                "canonical_source": canonical_rel,
                "canonical_file_fingerprints": fingerprints,
            },
            "summary": summary,
            "safety": safety,
            "top_products": top_products,
            "performance_coverage": top_metrics,
            "products": product_summaries,
            "named_subsets": subset_inventory,
            "warnings": warnings,
        }

        # Stage output is written to a temp sibling then atomically renamed.
        staging = store_root / "imports" / f".{dataset_id}.staging"
        if staging.exists():
            shutil.rmtree(staging)
        staging.mkdir(parents=True, exist_ok=True)

        _write_csv(
            staging / "samples.csv",
            ["dataset_id", "sample_key", "sample_name", "product_type", "craft"],
            samples_out,
        )
        _write_csv(
            staging / "formulas_long.csv",
            [
                "dataset_id", "sample_key", "sample_name",
                "product_type", "material_id", "amount",
            ],
            formulas_long,
        )
        _write_csv(
            staging / "performances_long.csv",
            [
                "dataset_id", "sample_key", "sample_name",
                "product_type", "metric", "value_raw",
                "value_numeric", "is_numeric",
            ],
            performances_long,
        )
        _write_csv(
            staging / "process_metadata.csv",
            [
                "dataset_id", "sample_key", "sample_name",
                "product_type", "LOGINCATEGORY", "TASKCATEGORY",
                "classification",
            ],
            process_out,
        )

        wide_fields = (
            [
                "dataset_id", "sample_key", "sample_name",
                "product_type", "craft",
                "meta::LOGINCATEGORY", "meta::TASKCATEGORY",
            ]
            + [f"formula::{x}" for x in material_ids]
            + [f"performance::{x}" for x in performance_metrics]
        )
        _write_csv(staging / "catalog_wide.csv", wide_fields, wide_rows)

        _atomic_json(staging / "products.json", {
            "products": product_summaries
        })
        _atomic_json(staging / "manifest.json", manifest)

        import_dir.parent.mkdir(parents=True, exist_ok=True)
        os.replace(staging, import_dir)

    _atomic_json(
        store_root / "current.json",
        {
            "dataset_id": dataset_id,
            "activated_at": utc_now_iso(),
            "source_sha256": archive_sha,
        },
    )

    return {
        "idempotent_replay": False,
        "manifest": manifest,
        "manifest_path": str(manifest_path),
    }
